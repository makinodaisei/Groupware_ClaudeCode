#!/usr/bin/env python3
"""
md2excel.py - Markdown table → Excel converter

docs/skills.md で定義された構造（Markdownテーブル）を解析し、
Excelファイルに変換する。

Usage:
    python tools/md2excel.py --input docs/db-definition.md --output artifacts/db-definition.xlsx
    python tools/md2excel.py --input docs/api-spec.md --output artifacts/api-spec.xlsx
    python tools/md2excel.py --all   # docs/*.md を一括変換

Requirements:
    pip install openpyxl pyyaml
"""

import argparse
import glob
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import yaml
except ImportError:
    print("ERROR: pyyaml is required. Run: pip install pyyaml", file=sys.stderr)
    sys.exit(1)


# ---------- Style constants ----------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ODD_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
EVEN_ROW_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
BORDER_SIDE = Side(style="thin", color="B0C4DE")
CELL_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE
)
SHEET_TITLE_FONT = Font(bold=True, size=14)


# ---------- Parsing ----------

def parse_frontmatter(content: str) -> tuple[dict, str]:
    """Extract YAML frontmatter from Markdown content. Returns (meta, body)."""
    if not content.startswith("---"):
        return {}, content

    end = content.find("\n---", 3)
    if end == -1:
        return {}, content

    yaml_str = content[3:end].strip()
    body = content[end + 4:].strip()
    try:
        meta = yaml.safe_load(yaml_str) or {}
    except yaml.YAMLError:
        meta = {}
    return meta, body


def parse_markdown_tables(content: str) -> list[tuple[str, list[str], list[list[str]]]]:
    """
    Parse all Markdown tables from content.
    Returns list of (sheet_name, headers, rows).
    sheet_name is derived from the nearest preceding H2/H3 heading.
    """
    results = []
    lines = content.splitlines()
    current_heading = "Sheet"
    heading_pattern = re.compile(r"^#{1,3}\s+(.+)$")
    separator_pattern = re.compile(r"^\|[\s|:-]+\|$")
    table_row_pattern = re.compile(r"^\|.+\|$")

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Track headings for sheet naming
        m = heading_pattern.match(line)
        if m:
            current_heading = m.group(1).strip()
            i += 1
            continue

        # Detect table header row
        if table_row_pattern.match(line) and i + 1 < len(lines):
            sep_line = lines[i + 1].strip()
            if separator_pattern.match(sep_line):
                # Parse header
                headers = _parse_row(line)

                # Parse body rows
                rows = []
                j = i + 2
                while j < len(lines):
                    row_line = lines[j].strip()
                    if table_row_pattern.match(row_line):
                        rows.append(_parse_row(row_line))
                        j += 1
                    else:
                        break

                results.append((current_heading, headers, rows))
                i = j
                continue

        i += 1

    return results


def _parse_row(line: str) -> list[str]:
    """Parse a Markdown table row into a list of cell values."""
    # Remove leading/trailing pipes and split
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    cells = [cell.strip() for cell in stripped.split("|")]
    return cells


# ---------- Excel generation ----------

def _auto_fit_column(ws, col_idx: int, values: list[str], max_width: int = 60):
    """Set column width based on max content length."""
    max_len = max((len(str(v)) for v in values), default=10)
    width = min(max_len + 4, max_width)
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_table_to_sheet(ws, sheet_title: str, headers: list[str], rows: list[list[str]]):
    """Write a single table to an Excel worksheet with styling."""
    # Sheet title row
    ws.append([sheet_title])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = SHEET_TITLE_FONT
    ws.row_dimensions[1].height = 20
    ws.append([])  # blank row

    # Header row
    header_row_idx = 3
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    ws.row_dimensions[header_row_idx].height = 20

    # Data rows
    for row_idx, row in enumerate(rows, start=4):
        # Pad row to match header length
        padded = row + [""] * (len(headers) - len(row))
        padded = padded[:len(headers)]
        ws.append(padded)

        fill = EVEN_ROW_FILL if (row_idx % 2 == 0) else ODD_ROW_FILL
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = CELL_BORDER
        ws.row_dimensions[row_idx].height = 18

    # Auto-fit columns based on all values
    for col_idx, header in enumerate(headers, start=1):
        col_values = [header] + [
            (row[col_idx - 1] if col_idx - 1 < len(row) else "")
            for row in rows
        ]
        _auto_fit_column(ws, col_idx, col_values)

    # Freeze panes at header row
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)


def _safe_sheet_name(name: str, existing: set) -> str:
    """Ensure sheet name is unique and within Excel's 31-char limit."""
    safe = re.sub(r'[\\/*?:\[\]]', "_", name)[:31]
    base = safe
    counter = 2
    while safe in existing:
        suffix = f"_{counter}"
        safe = base[:31 - len(suffix)] + suffix
        counter += 1
    return safe


def convert_md_to_excel(input_path: str, output_path: str) -> int:
    """
    Convert a Markdown file to an Excel workbook.
    Returns the number of sheets written.
    """
    content = Path(input_path).read_text(encoding="utf-8")
    meta, body = parse_frontmatter(content)
    tables = parse_markdown_tables(body)

    if not tables:
        print(f"WARNING: No tables found in {input_path}", file=sys.stderr)
        return 0

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    used_names = set()
    for sheet_heading, headers, rows in tables:
        if not headers:
            continue
        sheet_name = _safe_sheet_name(sheet_heading, used_names)
        used_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        write_table_to_sheet(ws, sheet_heading, headers, rows)
        print(f"  + Sheet: {sheet_name!r} ({len(rows)} rows)")

    # Ensure output directory exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")
    return len(used_names)


# ---------- CLI ----------

def main():
    parser = argparse.ArgumentParser(
        description="Convert Markdown tables to Excel (docs/skills.md rule-based)"
    )
    parser.add_argument("--input", "-i", help="Input Markdown file path")
    parser.add_argument("--output", "-o", help="Output Excel file path")
    parser.add_argument(
        "--all", "-a", action="store_true",
        help="Convert all docs/*.md files using their Frontmatter output_xlsx"
    )
    parser.add_argument(
        "--docs-dir", default="docs",
        help="Directory containing Markdown files (used with --all)"
    )
    args = parser.parse_args()

    if args.all:
        md_files = sorted(glob.glob(os.path.join(args.docs_dir, "*.md")))
        # Skip skills.md itself (it's the rules file, not a data doc)
        md_files = [f for f in md_files if not f.endswith("skills.md")]

        if not md_files:
            print(f"No .md files found in {args.docs_dir}/", file=sys.stderr)
            sys.exit(1)

        total_sheets = 0
        for md_path in md_files:
            content = Path(md_path).read_text(encoding="utf-8")
            meta, _ = parse_frontmatter(content)
            output_path = meta.get("output_xlsx")

            if not output_path:
                print(f"SKIP {md_path}: no output_xlsx in frontmatter", file=sys.stderr)
                continue

            print(f"\nProcessing: {md_path} → {output_path}")
            total_sheets += convert_md_to_excel(md_path, output_path)

        print(f"\nDone. Total sheets written: {total_sheets}")

    elif args.input and args.output:
        print(f"Processing: {args.input} → {args.output}")
        count = convert_md_to_excel(args.input, args.output)
        if count == 0:
            sys.exit(1)

    else:
        parser.print_help()
        print("\nExamples:")
        print("  python tools/md2excel.py --input docs/db-definition.md --output artifacts/db-definition.xlsx")
        print("  python tools/md2excel.py --all")
        sys.exit(1)


if __name__ == "__main__":
    main()
